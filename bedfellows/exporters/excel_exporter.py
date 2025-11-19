"""Excel export functionality."""

import logging
from datetime import datetime, date
from typing import Any, Optional
from decimal import Decimal

from peewee import ModelSelect

from bedfellows.exporters.base import BaseExporter

logger = logging.getLogger(__name__)


class ExcelExporter(BaseExporter):
    """Export data to Excel format (.xlsx)."""

    def __init__(self, output_path: str):
        """
        Initialize Excel exporter.

        Args:
            output_path: Path to output Excel file
        """
        super().__init__(output_path)
        self.validate_path(".xlsx")

        # Check if openpyxl is available
        try:
            import openpyxl
            self.openpyxl = openpyxl
        except ImportError:
            raise ImportError(
                "openpyxl is required for Excel export. "
                "Install with: pip install openpyxl"
            )

    def export(
        self,
        data: Any,
        model: Optional[type] = None,
        query: Optional[ModelSelect] = None,
    ) -> None:
        """
        Export data to Excel.

        Args:
            data: Data to export (list of dicts)
            model: Model class (optional)
            query: Query (optional)
        """
        if not data:
            logger.warning("No data to export")
            return

        # Ensure data is a list
        if not isinstance(data, list):
            data = [data]

        # Create workbook
        wb = self.openpyxl.Workbook()
        ws = wb.active

        # Set sheet name
        if model:
            ws.title = model._meta.table_name[:31]  # Excel limit is 31 chars
        else:
            ws.title = "Data"

        # Get field names from first record
        fieldnames = list(data[0].keys())

        # Write header row
        for col, fieldname in enumerate(fieldnames, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = fieldname
            cell.font = self.openpyxl.styles.Font(bold=True)

        # Write data rows
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, fieldname in enumerate(fieldnames, 1):
                value = row_data[fieldname]

                # Convert types for Excel
                if isinstance(value, (datetime, date)):
                    value = value
                elif isinstance(value, Decimal):
                    value = float(value)
                elif value is None:
                    value = ""

                ws.cell(row=row_idx, column=col_idx).value = value

        # Auto-size columns
        for column_cells in ws.columns:
            length = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = min(
                length + 2, 50
            )

        # Save workbook
        wb.save(self.output_path)

        logger.info(f"Exported {len(data)} records to {self.output_path}")

    def export_multiple(self, datasets: dict) -> None:
        """
        Export multiple datasets to separate sheets in one Excel file.

        Args:
            datasets: Dictionary of sheet_name -> data mappings
        """
        wb = self.openpyxl.Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        for sheet_name, data in datasets.items():
            if not data:
                continue

            # Ensure data is a list
            if not isinstance(data, list):
                data = [data]

            # Create sheet (limit name to 31 chars for Excel)
            ws = wb.create_sheet(title=sheet_name[:31])

            # Get field names
            fieldnames = list(data[0].keys())

            # Write header
            for col, fieldname in enumerate(fieldnames, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = fieldname
                cell.font = self.openpyxl.styles.Font(bold=True)

            # Write data
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, fieldname in enumerate(fieldnames, 1):
                    value = row_data[fieldname]

                    if isinstance(value, (datetime, date)):
                        value = value
                    elif isinstance(value, Decimal):
                        value = float(value)
                    elif value is None:
                        value = ""

                    ws.cell(row=row_idx, column=col_idx).value = value

            # Auto-size columns
            for column_cells in ws.columns:
                length = max(len(str(cell.value or "")) for cell in column_cells)
                ws.column_dimensions[column_cells[0].column_letter].width = min(
                    length + 2, 50
                )

        # Save workbook
        wb.save(self.output_path)

        total_records = sum(len(d) if isinstance(d, list) else 1 for d in datasets.values())
        logger.info(
            f"Exported {total_records} records across {len(datasets)} sheets to {self.output_path}"
        )
